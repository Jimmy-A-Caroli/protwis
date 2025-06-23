from django.shortcuts import render
from django.http import HttpResponse, HttpResponseBadRequest
from django.conf import settings
from django.db.models import Q
from django.views.generic import TemplateView
from django import forms
from protein.models import Protein, ProteinFamily
from common.phylogenetic_tree import PhylogeneticTreeGenerator

import json
from copy import deepcopy
from collections import OrderedDict
import sys


try:
    import importlib.metadata
except ImportError:
    sys.modules['importlib.metadata'] = __import__('importlib_metadata')

import pandas as pd
import numpy as np
from sklearn.manifold import TSNE
from sklearn.cluster import KMeans
import matplotlib.colors as mcolors
import re

import openpyxl
import os


class DataMapperHome(TemplateView):

    def dispatch(self, request, *args, **kwargs):
        """Ensure 'page' is always set for both GET and POST requests"""

        # Extract from URL (GET)
        if 'page' in kwargs:
            self.page = kwargs['page']

        # Extract from POST (ensure it is there)
        elif request.method == 'POST':
            self.page = request.POST.get('page')

        # If 'page' is still None, return an error
        if not self.page:
            return HttpResponseBadRequest("Missing 'page' parameter")

        return super().dispatch(request, *args, **kwargs)

    def get_template_names(self):
        """Ensure the correct template is used"""
        return [f'mapper/DataMapperHome{self.page}.html']

    def get_context_data(self, **kwargs):
        """Pass the page variable into the context"""
        context = super().get_context_data(**kwargs)
        context['page'] = self.page  # Pass 'page' to the template
        return context

    @staticmethod
    def keep_by_names(data, names_to_keep):
        data_copy = deepcopy(data)
        if isinstance(data_copy, list):
            # Process each item in the list
            kept_items = [DataMapperHome.keep_by_names(item, names_to_keep) for item in data_copy]
            # Return only non-None items
            return [item for item in kept_items if item is not None]
        elif isinstance(data_copy, OrderedDict):
            name = data_copy.get('name')
            if name not in names_to_keep.keys():
                if 'children' in data_copy:
                    # Recursively process children
                    data_copy['children'] = DataMapperHome.keep_by_names(data_copy['children'], names_to_keep)
                    # Remove the 'children' key if it's empty after processing
                    if not data_copy['children']:
                        return None
                else:
                    return None
            else:
                # If the name is in the keep list, update the 'value' field
                if 'Inner' in names_to_keep[name]:
                    data_copy['value'] = names_to_keep[name]['Inner']
                # Process children if present
                if 'children' in data_copy:
                    data_copy['children'] = DataMapperHome.keep_by_names(data_copy['children'], names_to_keep)
                    if not data_copy['children']:
                        del data_copy['children']
            return data_copy
        return data_copy

    @staticmethod
    def convert_keys(datatree, conversion):
        new_tree = {}
        for key, value in datatree.items():
            # Convert the key using the conversion dictionary
            new_key = conversion.get(key, key)  # Fallback to the original key if no conversion is found
            if isinstance(value, dict):
                # Recursively convert keys of nested dictionaries
                new_tree[new_key] = DataMapperHome.convert_keys(value, conversion)
            else:
                # If the value is a list (end of the branch), just assign it
                new_tree[new_key] = value
        return new_tree

    @staticmethod
    def reassign_keys(original_dict, key_mappings):
         # Create a new dictionary to store the reassigned keys
         new_dict = {}
         # Iterate over the list of tuples (new_key, old_key)
         for new_key, old_key in key_mappings:
             # Check if the old_key exists in the original dictionary
             if old_key in original_dict:
                 # Assign the value from the original dictionary to the new key
                 new_dict[new_key] = original_dict[old_key]

         return new_dict

    @staticmethod
    def filter_and_extend_dict(big_dict, small_dict):
        def filter_dict(d, keys_set):
            filtered_dict = {}
            for key, value in d.items():
                if isinstance(value, dict):
                    filtered_sub_dict = filter_dict(value, keys_set)
                    if filtered_sub_dict:
                        filtered_dict[key] = filtered_sub_dict
                elif isinstance(value, list):
                    filtered_values = [v for v in value if v in keys_set]
                    if filtered_values:
                        filtered_dict[key] = filtered_values
            return filtered_dict

        def extend_dict(d, additional_values):
            for key, value in d.items():
                if isinstance(value, dict):
                    extend_dict(value, additional_values)
                elif isinstance(value, list):
                    extended_list = []
                    for item in value:
                        if item in additional_values:
                            # Wrap values into a list and add an extra dictionary layer
                            extended_list.append({item: [additional_values[item]['Value1'], additional_values[item]['Value2']]})
                        else:
                            extended_list.append(item)
                    d[key] = extended_list

        # Step 1: Filter the big dictionary
        filtered_dict = filter_dict(big_dict, set(small_dict.keys()))

        # Step 2: Extend the filtered dictionary
        extend_dict(filtered_dict, small_dict)

        return filtered_dict

    @staticmethod
    def filter_dict(d, elements):
        filtered_dict = {}
        for key, value in d.items():
            if isinstance(value, dict):
                filtered_sub_dict = DataMapperHome.filter_dict(value, elements)
                if filtered_sub_dict:
                    filtered_dict[key] = filtered_sub_dict
            elif isinstance(value, list):
                filtered_values = [v for v in value if v in elements]
                if filtered_values:
                    filtered_dict[key] = filtered_values
        return filtered_dict

    @staticmethod
    def generate_list_plot(listplot): #ADD AN INPUT FILTER DICTIONARY
        # Generate the master dict of protein families

        data = list(listplot.keys())
        names = list(Protein.objects.filter(entry_name__in=data).values_list('name', flat=True))
        # Names conversion dict
        names_dict = Protein.objects.filter(entry_name__in=data).values('entry_name', 'name').order_by('entry_name')
        names_conversion_dict = {item['entry_name']: item['name'] for item in names_dict}
        IUPHAR_to_uniprot_dict = {item['name']: item['entry_name'] for item in names_dict}

        families = ProteinFamily.objects.all()
        datatree = {}
        conversion = {}
        for item in families:
            if len(item.slug) == 3 and item.slug not in datatree.keys():
                datatree[item.slug] = {}
                conversion[item.slug] = item.name
            if len(item.slug) == 7 and item.slug not in datatree[item.slug[:3]].keys():
                datatree[item.slug[:3]][item.slug[:7]] = {}
                conversion[item.slug] = item.name
            if len(item.slug) == 11 and item.slug not in datatree[item.slug[:3]][item.slug[:7]].keys():
                datatree[item.slug[:3]][item.slug[:7]][item.slug[:11]] = []
                conversion[item.slug] = item.name
            if len(item.slug) == 15 and item.slug not in datatree[item.slug[:3]][item.slug[:7]][item.slug[:11]]:
                datatree[item.slug[:3]][item.slug[:7]][item.slug[:11]].append(item.name)

        datatree2 = DataMapperHome.convert_keys(datatree, conversion)
        datatree2.pop('Parent family', None)
        datatree3 = DataMapperHome.filter_dict(datatree2, names)
        data_converted = {names_conversion_dict[key]: value for key, value in listplot.items()}
        Data_full = {"NameList": datatree3, "DataPoints": data_converted, "LabelConversionDict":IUPHAR_to_uniprot_dict}
        return Data_full

    @staticmethod
    def GenerateGPCRomeDataStructure(data_type: str = "Classic"):

        # Determine which proteins and families to use based on type
        if data_type == "Classic":
            all_proteins = Protein.objects.filter(
                species_id=1,
                parent_id__isnull=True,
                accession__isnull=False,
                family_id__slug__startswith='0'
            ).exclude(
                Q(family_id__slug__startswith='007') |
                Q(family_id__slug__startswith='008')
            )

            families = ProteinFamily.objects.exclude(slug='000')

            # Get valid names (for pruning) from all_proteins
            valid_names = set(all_proteins.values_list('name', flat=True))

            # Build name-to-entry_name mapping
            proteins = Protein.objects.filter(
                species_id=1,
                name__in=valid_names
            ).values(
                'entry_name', 'name'
            ).order_by('entry_name')

            name_to_entry = {item['name']: item['entry_name'] for item in proteins}

            # Create flat list of entry names
            entry_names = [item['entry_name'] for item in proteins]

            # Fetch proteins with prefetch on genes
            protein_genes = Protein.objects.prefetch_related('genes').filter(entry_name__in=entry_names)

            # Create mapping from entry_name to gene name (position == 0 only)
            entry_to_gene = {
                protein.entry_name: next((g.name for g in protein.genes.all() if g.position == 0), None)
                for protein in protein_genes
            }

        elif data_type == "Odorant":
            all_proteins = Protein.objects.filter(
                species_id=1,
                parent_id__isnull=True,
                accession__isnull=False
            ).filter(
                Q(family_id__slug__startswith='007') | Q(family_id__slug__startswith='008')
            )

            families = ProteinFamily.objects.filter(
                Q(slug__startswith='007') | Q(slug__startswith='008')
            )

            valid_names = set(all_proteins.values_list('name', flat=True))

            proteins = Protein.objects.filter(
                species_id=1,
                name__in=valid_names
            ).values(
                'entry_name', 'name'
            ).order_by('entry_name')

            name_to_entry = {item['name']: item['entry_name'] for item in proteins}

            entry_names = [item['entry_name'] for item in proteins]

            # Fetch proteins with prefetch on genes
            protein_genes = Protein.objects.prefetch_related('genes').filter(entry_name__in=entry_names)

            # Create mapping from entry_name to gene name (position == 0 only)
            entry_to_gene = {
                protein.entry_name: next((g.name for g in protein.genes.all() if g.position == 0), None)
                for protein in protein_genes
            }

        else:
            raise ValueError(f"Unsupported data structure type: {type}")

        # Tree building
        def build_family_tree(families):
            datatree = {}
            slug_to_name = {fam.slug: fam.name for fam in families}

            for item in families:
                slug_parts = item.slug.split('_')
                current_level = datatree

                for i in enumerate(slug_parts):
                    full_slug = '_'.join(slug_parts[:i + 1])
                    name = item.name if full_slug == item.slug else None

                    if i == len(slug_parts) - 1:
                        if len(slug_parts) == 4:
                            if item.name in name_to_entry:
                                entry_name = name_to_entry[item.name]
                                entry_code = entry_name.split('_')[0].upper()
                                gene_symbol = entry_to_gene.get(entry_name)  # Entrez name, if available
                                current_level[item.name] = {
                                    "Data": "Empty",
                                    "EntryName": entry_code,
                                    "Entrez": gene_symbol if gene_symbol else "UNKNOWN",
                                    "Color": "#FFFFFF"
                                }
                        else:
                            current_level.setdefault(name, {})
                    else:
                        if name:
                            current_level = current_level.setdefault(name, {})
                        else:
                            current_level = current_level.setdefault(full_slug, {})  # Temp key

            def convert_keys(tree):
                new_tree = {}
                for key, value in tree.items():
                    new_key = slug_to_name.get(key, key)
                    new_tree[new_key] = convert_keys(value) if isinstance(value, dict) else value
                return new_tree

            return convert_keys(datatree)

        # Prune tree to remove anything not in valid_names
        def prune_tree(tree, valid_leaves):
            pruned = {}
            for key, value in tree.items():
                if isinstance(value, dict):
                    if 'Data' in value:
                        if key in valid_leaves:
                            pruned[key] = value
                    else:
                        pruned_subtree = prune_tree(value, valid_leaves)
                        if pruned_subtree:
                            pruned[key] = pruned_subtree
            return pruned if pruned else None

        # Build and prune the tree
        datatree = build_family_tree(families)
        GPCRomeStructureDict = prune_tree(datatree, valid_names)

        # Class renaming
        class_rename_map = {
            "Class A (Rhodopsin)": "A",
            "Class B1 (Secretin)": "B1",
            "Class B2 (Adhesion)": "B2",
            "Class C (Glutamate)": "C",
            "Class F (Frizzled)": "F",
            "Class O1 (fish-like odorant)": "O1",
            "Class O2 (tetrapod specific odorant)": "O2",
            "Class T2 (Taste 2)": "T2",
            "Other GPCRs": "Classless"
        }

        if data_type == "Classic" and GPCRomeStructureDict:
            GPCRome_dict = {
                "Circle_1": {},
                "Circle_2": {},
                "Circle_3": {},
                "Circle_4": {},
                "Circle_5": {}
            }

            class_A_receptor_families = 0

            for Class, ligand_types in GPCRomeStructureDict.items():
                renamed_class = class_rename_map.get(Class, Class)

                if Class == "Class A (Rhodopsin)":
                    sorted_receptor_families = []

                    for Ligand_type, receptor_families in ligand_types.items():
                        if Ligand_type == "Orphan receptors":
                            continue

                        for Receptor_Family in receptor_families:
                            if Receptor_Family == "Class A Orphans":
                                continue

                            sorted_receptor_families.append(
                                (Ligand_type, Receptor_Family, receptor_families[Receptor_Family])
                            )

                    sorted_receptor_families.sort(key=lambda x: x[1])  # sort by receptor family name

                    for Ligand_type, Receptor_Family, receptors in sorted_receptor_families:
                        target_circle = "Circle_1" if class_A_receptor_families < 43 else "Circle_2"

                        GPCRome_dict.setdefault(target_circle, {}).setdefault(renamed_class, {}).setdefault(Ligand_type, {})[Receptor_Family] = receptors
                        class_A_receptor_families += 1

                    # Add Class A orphans to Circle_2 if present
                    orphans = ligand_types.get("Orphan receptors", {})
                    if "Class A orphans" in orphans:
                        GPCRome_dict["Circle_2"].setdefault(renamed_class, {}).setdefault("Orphan receptors", {})["Class A orphans"] = orphans["Class A orphans"]

                elif Class in ["Class B1 (Secretin)", "Class B2 (Adhesion)"]:
                    GPCRome_dict["Circle_3"].setdefault(renamed_class, {}).update(ligand_types)

                elif Class in ["Class C (Glutamate)", "Class F (Frizzled)"]:
                    GPCRome_dict["Circle_4"].setdefault(renamed_class, {}).update(ligand_types)

                elif Class in ["Class T2 (Taste 2)", "Other GPCRs"]:
                    GPCRome_dict["Circle_5"].setdefault(renamed_class, {}).update(ligand_types)

            # Remove the ligand type layer
            for circle in GPCRome_dict:
                for class_name in list(GPCRome_dict[circle].keys()):
                    new_structure = {}

                    for ligand_type in list(GPCRome_dict[circle][class_name].keys()):
                        for receptor_family, receptors in GPCRome_dict[circle][class_name][ligand_type].items():
                            new_structure[receptor_family] = receptors

                    # Replace the old structure with the flattened one
                    GPCRome_dict[circle][class_name] = new_structure
            # Final return (renamed and sorted into circles)
            return {
                "Data": GPCRome_dict
            }
        if data_type == "Odorant" and GPCRomeStructureDict:
            GPCRome_dict = {
                "Circle_1": {},  # Family 1–4 from Class O2
                "Circle_2": {},  # Family 5–9 from Class O2
                "Circle_3": {},  # Family 10–14 from Class O2
                "Circle_4": {}   # All of Class O1
            }

            for Class, ligand_types in GPCRomeStructureDict.items():
                renamed_class = class_rename_map.get(Class, Class)

                if Class == "Class O2 (tetrapod specific odorant)":
                    sorted_families = []

                    for Ligand_type, receptor_families in ligand_types.items():
                        for Receptor_Family, receptors in receptor_families.items():
                            try:
                                family_number = int(Receptor_Family.replace("Odorant family", "").strip())
                            except ValueError:
                                continue

                            sorted_families.append(
                                (family_number, Ligand_type, Receptor_Family, receptors)
                            )

                    sorted_families.sort(key=lambda x: x[0])  # Sort numerically by family number

                    for family_number, Ligand_type, Receptor_Family, receptors in sorted_families:
                        renamed_family = f"Family {family_number}"
                        if 1 <= family_number <= 4:
                            circle = "Circle_1"
                        elif 5 <= family_number <= 9:
                            circle = "Circle_2"
                        elif 10 <= family_number <= 14:
                            circle = "Circle_3"
                        else:
                            continue  # Skip any outside defined ranges

                        GPCRome_dict.setdefault(circle, {}).setdefault(renamed_class, {}).setdefault(Ligand_type, {})[renamed_family] = receptors

                elif Class == "Class O1 (fish-like odorant)":
                    renamed_ligand_types = {}

                    for Ligand_type, receptor_families in ligand_types.items():
                        renamed_receptor_families = {}

                        for Receptor_Family, receptors in receptor_families.items():
                            try:
                                family_number = int(Receptor_Family.replace("Odorant family", "").strip())
                                renamed_family = f"Family {family_number}"
                            except ValueError:
                                renamed_family = Receptor_Family  # fallback to original if parsing fails

                            renamed_receptor_families[renamed_family] = receptors

                        renamed_ligand_types[Ligand_type] = renamed_receptor_families

                    GPCRome_dict["Circle_4"].setdefault(renamed_class, {}).update(renamed_ligand_types)

            # Flatten ligand_type layer (remove ligand type level)
            for circle in GPCRome_dict:
                for class_name in list(GPCRome_dict[circle].keys()):
                    new_structure = {}
                    for ligand_type in GPCRome_dict[circle][class_name]:
                        for receptor_family, receptors in GPCRome_dict[circle][class_name][ligand_type].items():
                            new_structure[receptor_family] = receptors
                    GPCRome_dict[circle][class_name] = new_structure

            return {
                "Data": GPCRome_dict
            }

    @staticmethod
    def update_nested_GPCRome_data(structure_dict, raw_data):
        def normalize_key(raw_key):
            return raw_key.split("_")[0].upper()

        # Normalize and extract only non-null values
        normalized_raw_data = {
            normalize_key(key): {
                'Data': val.get("Value1"),
                'Color': val.get("Value2")  # Might be None
            }
            for key, val in raw_data.items()
            if isinstance(val, dict) and "Value1" in val
        }

        def recursive_update(d):
            if isinstance(d, dict):
                if "EntryName" in d and "Data" in d:
                    entry = d["EntryName"].upper()
                    if entry in normalized_raw_data:
                        d["Data"] = normalized_raw_data[entry]["Data"]
                        
                        color = normalized_raw_data[entry].get("Color")
                        if color is not None:
                            d["Color"] = color

                for value in d.values():
                    recursive_update(value)
            elif isinstance(d, list):
                for item in d:
                    recursive_update(item)

        recursive_update(structure_dict)
        return structure_dict

    @staticmethod
    def generate_tree_plot(input_data): #ADD AN INPUT FILTER DICTIONARY
        ### TREE SECTION
        tree = PhylogeneticTreeGenerator()
        class_a_data = tree.get_tree_data(ProteinFamily.objects.get(name='Class A (Rhodopsin)'))
        class_b1_data = tree.get_tree_data(ProteinFamily.objects.get(name__startswith='Class B1 (Secretin)'))
        class_b2_data = tree.get_tree_data(ProteinFamily.objects.get(name__startswith='Class B2 (Adhesion)'))
        class_c_data = tree.get_tree_data(ProteinFamily.objects.get(name__startswith='Class C (Glutamate)'))
        class_f_data = tree.get_tree_data(ProteinFamily.objects.get(name__startswith='Class F (Frizzled)'))
        class_t2_data = tree.get_tree_data(ProteinFamily.objects.get(name__startswith='Class T2 (Taste 2)'))
        class_cl_data = tree.get_tree_data(ProteinFamily.objects.get(name__startswith='Other GPCRs'))
        ### GETTING NODES
        data_a = class_a_data.get_nodes_dict(None)
        data_b1 = class_b1_data.get_nodes_dict(None)
        data_b2 = class_b2_data.get_nodes_dict(None)
        data_c = class_c_data.get_nodes_dict(None)
        data_f = class_f_data.get_nodes_dict(None)
        data_t2 = class_t2_data.get_nodes_dict(None)
        data_cl = class_cl_data.get_nodes_dict(None)
        #Collating everything into a single tree
        general_options = {'depth': 4,
                           'branch_length': {1: 'Class A (Rhodopsin)',
                                             2: 'Alicarboxylic acid',
                                             3: 'Gonadotrophin-releasing hormone',
                                             4: ''},
                           'branch_trunc': 0,
                           'leaf_offset': 30,
                           'anchor': "tree_plot",
                           'label_free': [],
                           'fontSize': {
                                'class': "15px",
                                'ligandtype': "14px",
                                'receptorfamily': "13px",
                                'receptor': "12px"
                            }}
        master_dict = OrderedDict([('name', ''),
                                   ('value', 3000),
                                   ('color', ''),
                                   ('children',[])])
        class_a_dict = OrderedDict([('name', 'Class A (Rhodopsin)'),
                                   ('value', 0),
                                   ('color', 'Red'),
                                   ('children',data_a['children'])])
        class_b1_dict = OrderedDict([('name', 'Class B1 (Secretin)'),
                                   ('value', 0),
                                   ('color', 'Green'),
                                   ('children',data_b1['children'])])
        class_b2_dict = OrderedDict([('name', 'Class B2 (Adhesion)'),
                                  ('value', 0),
                                  ('color', 'Blue'),
                                  ('children',data_b2['children'])])
        class_c_dict = OrderedDict([('name', 'Class C (Glutamate)'),
                                  ('value', 0),
                                  ('color', 'Purple'),
                                  ('children',data_c['children'])])
        class_f_dict = OrderedDict([('name', 'Class F (Frizzled)'),
                                  ('value', 0),
                                  ('color', 'Grey'),
                                  ('children',data_f['children'])])
        class_t2_dict = OrderedDict([('name', 'Class T2 (Taste 2)'),
                                  ('value', 0),
                                  ('color', 'Orange'),
                                  ('children',data_t2['children'])])
        class_cl_dict = OrderedDict([('name', 'Classless'),
                                  ('value', 0),
                                  ('color', 'Gold'),
                                  ('children',data_cl['children'])])
        ### APPENDING TO MASTER DICT
        master_dict['children'].append(class_a_dict)
        master_dict['children'].append(class_b1_dict)
        master_dict['children'].append(class_b2_dict)
        master_dict['children'].append(class_c_dict)
        master_dict['children'].append(class_f_dict)
        master_dict['children'].append(class_t2_dict)
        master_dict['children'].append(class_cl_dict)

        updated_data = {key.replace('_human', ''): value for key, value in input_data.items()}
        circles = {key.replace('_human', '').upper(): {k: v for k, v in value.items()} for key, value in input_data.items()}
        master_dict = DataMapperHome.keep_by_names(master_dict, updated_data)

        if len(master_dict['children']) == 1:
            master_dict = master_dict['children'][0]
            general_options['depth'] = 3
            general_options['branch_length'] = {1: 'Alicarboxylic acid',
                                             2: 'Gonadotrophin-releasing hormone',
                                             3: ''}
        else:
            pass

        entry_names_list = list(input_data.keys())  # or: input_data.keys()

        # Step 2: Trim protein queries
        whole_receptors = Protein.objects.filter(entry_name__in=entry_names_list).prefetch_related("family", "family__parent__parent__parent")
            
        protein_genes = Protein.objects.filter(entry_name__in=entry_names_list).prefetch_related('genes')
            
        entry_to_gene = {
            protein.entry_name: next((g.name for g in protein.genes.all() if g.position == 0), None)
            for protein in protein_genes
        }

        whole_rec_dict = {}
        entrez_label_dict = {}
        for rec in whole_receptors:
            rec_entryname = rec.entry_name
            rec_uniprot = rec.entry_short()
            rec_iuphar = rec.family.name.replace("receptor", '').replace("<i>", "").replace("</i>", "").strip()
            if (rec_iuphar[0].isupper()) or (rec_iuphar[0].isdigit()):
                whole_rec_dict[rec_uniprot] = [rec_iuphar]
            else:
                whole_rec_dict[rec_uniprot] = [rec_iuphar.capitalize()]
            # Entrez label — safely
            gene_name = entry_to_gene.get(rec_entryname)
            if gene_name:
                entrez_label_dict[rec_uniprot] = [gene_name]

        return master_dict, general_options, circles, whole_rec_dict, entrez_label_dict

    @staticmethod
    def clustering_test(method, data, data_type):
        # Convert the nested dictionary to a DataFrame
        data = {key.replace('_human', ''): value for key, value in data.items()}
        data_df = pd.DataFrame(data).T

        # Ensure Value1 is always filled with 0 if missing
        if 'Value1' in data_df.columns:
            data_df['Value1'] = data_df['Value1'].fillna(0)

        if 'Value2' in data_df.columns:
            if data_df['Value2'].isnull().sum() > 0:
                data_df['Value2'].fillna(data_df['Value2'].mean(), inplace=True)

            # --- Compute distance matrix from Value2 ---
            values = data_df[['Value2']].values  # shape: (n, 1)
            distance_matrix = np.abs(values - values.T)  # shape: (n, n)
            distance_df = pd.DataFrame(distance_matrix, index=data_df.index, columns=data_df.index)

            # --- Run reduction ---
            reduced_df = DataMapperHome.reduce_and_cluster(distance_df, method=method, is_distance_matrix=True)

            # --- Merge metadata ---
            df_merged = pd.merge(reduced_df, data_df['Value1'], left_on='label', right_index=True, how='left')
            df_merged.rename(columns={'Value1': 'fill'}, inplace=True)

            ## add class/ligand_type/receptor_family clusters ##

            # Step 1: Fetch data
            proteins = Protein.objects.filter(
                parent_id__isnull=True, species_id=1
            ).values_list(
                'entry_name',
                "family__parent__parent__parent__name",  # To be renamed as 'Class'
                'family__parent__parent__name',  # To be renamed as 'Ligand type'
                'family__parent__name'  # To be renamed as 'Receptor family'
            )

            # Step 2: Convert to a DataFrame
            proteins_df = pd.DataFrame(list(proteins), columns=['entry_name', 'Class', 'Ligand type', 'Receptor family'])

            # Step 3: Remove '_human' suffix from 'entry_name'
            proteins_df['entry_name'] = proteins_df['entry_name'].str.replace('_human', '')

            # Step 4: Rename 'entry_name' to 'label'
            proteins_df = proteins_df.rename(columns={'entry_name': 'label'})

            # Step 5: Merge with reduced_df on 'label'
            merged_df = pd.merge(df_merged, proteins_df, on='label', how='left')
            # Prepare the data for visualization
            data_json = merged_df.to_json(orient='records')
        else:
            if data_type == 'seq':
                # Get the info of the plot
                full_matrix = DataMapperHome.generate_full_matrix(method)
                # full_matrix_structure = DataMapperHome.generate_full_matrix_structure(method)

                # Filter the original fill matrix based on what we use provided
                reduced_input = full_matrix[full_matrix['label'].isin(list(data.keys()))]
                data_df = pd.DataFrame(data).T
                # Merge the dataframes
                df_merged = pd.merge(reduced_input, data_df['Value1'], left_on='label', right_index=True, how='left')
                df_merged.rename(columns={'Value1': 'fill'}, inplace=True)
                # Prepare the data for visualization
                data_json = df_merged.to_json(orient='records')
            elif data_type == 'structure':
                # Get the info of the plot
                full_matrix_structure = DataMapperHome.generate_full_matrix_structure(method)

                # Filter the original fill matrix based on what we use provided
                reduced_input = full_matrix_structure[full_matrix_structure['label'].isin(list(data.keys()))]
                data_df = pd.DataFrame(data).T
                # Merge the dataframes
                df_merged = pd.merge(reduced_input, data_df['Value1'], left_on='label', right_index=True, how='left')
                df_merged.rename(columns={'Value1': 'fill'}, inplace=True)
                # Prepare the data for visualization
                data_json = df_merged.to_json(orient='records')

        return data_json

    # Generate full similarity matrix for cluster or load existing #
    def generate_full_matrix(method):
        Data_dir = settings.DATA_DIR
        output_file = os.sep.join([Data_dir, 'structure_data', 'HumanGPCRSimilarityAllData_{}.csv'.format(method)])
        # Check if the file exists
        if os.path.exists(output_file):
            # Load the data from the existing file
            merged_df = pd.read_csv(output_file, index_col=0)
        else:
            # Original processing steps
            similarity_matrix_file = os.sep.join([Data_dir, 'structure_data', 'human_gpcr_similarity_data_all_segments.csv'])
            data = pd.read_csv(similarity_matrix_file)
            data = data[['receptor1_entry_name', 'receptor2_entry_name', 'similarity']]
            matrix = data.pivot(index='receptor1_entry_name', columns='receptor2_entry_name', values='similarity')
            matrix.index = matrix.index.str.replace('_human', '', regex=False)
            matrix.columns = matrix.columns.str.replace('_human', '', regex=False)
            matrix = matrix.fillna(100)

            # ---------------------------------------------
            # Step 2: Normalize Similarity Values to [0, 1]
            # ---------------------------------------------

            normalized_matrix = matrix / 100.0

            distance_matrix_df = 1.0 - normalized_matrix

            # Perform reduction and clustering
            reduced_df = DataMapperHome.reduce_and_cluster(distance_matrix_df, method='tsne')

            reduced_df['label'] = reduced_df['label'].apply(lambda x: x.split('[Human] ')[1] if '[Human] ' in x else x)
            reduced_df['label'] = reduced_df['label'].apply(lambda x: x.split('_human')[0] if '_human' in x else x)

            # add class/ligand_type/receptor_family clusters

            # Step 1: Fetch data
            proteins = Protein.objects.filter(
                parent_id__isnull=True,species_id=1
            ).values_list(
                'entry_name',
                "family__parent__parent__parent__name",  # To be renamed as 'Class'
                'family__parent__parent__name',  # To be renamed as 'Ligand type'
                'family__parent__name'  # To be renamed as 'Receptor family'
            )

            # Step 2: Convert to a DataFrame
            proteins_df = pd.DataFrame(list(proteins), columns=['entry_name', 'Class', 'Ligand type', 'Receptor family'])
            proteins_df.to_excel(os.sep.join([settings.DATA_DIR, 'structure_data', 'All_GPCRs_ligandType_Families.xlsx']),index=False)

            # Step 3: Remove '_human' suffix from 'entry_name'
            proteins_df['entry_name'] = proteins_df['entry_name'].str.replace('_human', '')

            # Step 4: Rename 'entry_name' to 'label'
            proteins_df = proteins_df.rename(columns={'entry_name': 'label'})

            # Step 5: Merge with reduced_df on 'label'
            merged_df = pd.merge(reduced_df, proteins_df, on='label', how='left')

            # Save the reduced DataFrame to a CSV file
            merged_df.to_csv(output_file)

        return merged_df

    @staticmethod
    def reduce_and_cluster(data, method='tsne', n_components=2, n_clusters=5, is_distance_matrix=False):

        # Figure out how many points we have
        n_points = data.shape[0]

        # Simple heuristic: scale perplexity and clamp it between 2 and 50
        suggested_perplexity = max(2, min(n_points / 5, 50))

        if method == 'tsne':
            if is_distance_matrix:
                reducer = TSNE(n_components=n_components, metric='precomputed', random_state=42,perplexity=suggested_perplexity)
            else:
                reducer = TSNE(n_components=n_components, random_state=42,perplexity=suggested_perplexity)
        else:
            raise ValueError("Method should be 'tsne' for now (or add other methods)")

        reduced_data = reducer.fit_transform(data)

        kmeans = KMeans(n_clusters=n_clusters, random_state=42)
        clusters = kmeans.fit_predict(reduced_data)

        df = pd.DataFrame(reduced_data, columns=['x', 'y'])
        df['cluster'] = clusters
        df['label'] = data.index if hasattr(data, 'index') else range(len(data))

        return df

    @staticmethod
    def map_to_quartile(value, quartiles):
        if value <= quartiles[0.25]:
            return 10
        elif value <= quartiles[0.5]:
            return 20
        elif value <= quartiles[0.75]:
            return 30
        else:
            return 40

    @staticmethod
    def Label_conversion_info(data):
        # Get list of keys (UniProt-style entry names)
        Name_list = list(data.keys())

        # Fetch IUPHAR names
        names_dict = Protein.objects.filter(entry_name__in=Name_list).values('entry_name', 'name').order_by('entry_name')
        UniProt_to_IUPHAR_converter = {item['entry_name']: item['name'] for item in names_dict}
        IUPHAR_to_UniProt_converter = {item['name']: item['entry_name'] for item in names_dict}

        # Fetch genes with position == 0
        protein_genes = Protein.objects.prefetch_related('genes').filter(entry_name__in=Name_list)
        UniProt_to_Gene_converter = {
            protein.entry_name: next((g.name for g in protein.genes.all() if g.position == 0), None)
            for protein in protein_genes
        }

        # Create IUPHAR to Gene mapping using the two converters above
        IUPHAR_to_Gene_converter = {
            iuphar: UniProt_to_Gene_converter.get(entry)
            for iuphar, entry in IUPHAR_to_UniProt_converter.items()
            if UniProt_to_Gene_converter.get(entry) is not None
        }

        # Final converter dict
        Label_converter = {
            'UniProt_to_IUPHAR_converter': UniProt_to_IUPHAR_converter,
            'IUPHAR_to_UniProt_converter': IUPHAR_to_UniProt_converter,
            'UniProt_to_Gene_converter': UniProt_to_Gene_converter,
            'IUPHAR_to_Gene_converter': IUPHAR_to_Gene_converter
        }

        return Label_converter
    

    def post(self, request, *args, **kwargs):
        ### This method handles POST requests for form submission ###

        if request.method == 'POST':

            # Utilize ExcelUploadForm class #
            form = ExcelUploadForm(request.POST,request.FILES)

            # If form is valid #
            if form.is_valid():

                # Get cleaned data #
                file = form.cleaned_data['file']

                # Check if file is .xlsx #
                if not file.name.endswith('.xlsx'):
                    return render(request, f'mapper/DataMapperHome{self.page}.html', {'upload_status': 'Failed','Error_message': "The uploaded file is not an .xlsx file."})
                else:
                    try:
                        workbook = openpyxl.load_workbook(filename=file,read_only=False)
                    except:
                        return render(request, f'mapper/DataMapperHome{self.page}.html', {'upload_status': 'Failed','Error_message': "Unable to load excel file, might be corrupted or not inline with the template file."})

                    if workbook:

                        # Fetch protein data for processing 1 (Cluster, List, and Heatmap) #
                        all_proteins = Protein.objects.filter(species_id=1, parent_id__isnull=True, accession__isnull=False,family_id__slug__startswith='0').values_list('entry_name', flat=True).distinct()

                        # Fetch protein data for processing 2 (GPCRome wheel and Tree plot) #
                        Proteins_GPCRomeTree = Protein.objects.filter(
                                species_id=1, 
                                parent_id__isnull=True,
                                accession__isnull=False,
                                family_id__slug__startswith='0'
                            ).exclude(
                                family_id__slug__startswith='007'
                            ).exclude(
                                family_id__slug__startswith='008'
                            ).values_list('entry_name', flat=True).distinct()
                        
                        # Fetch proteins with prefetch on genes
                        proteins = Protein.objects.prefetch_related('genes').filter(entry_name__in=all_proteins)

                        # Create mapping from entry_name to gene name (position == 0 only)
                        entry_to_gene = {
                            protein.entry_name: next((g.name for g in protein.genes.all() if g.position == 0), None)
                            for protein in proteins
                        }

                        # Reverse gene → entry_name mapping
                        gene_to_entry = {v.upper(): k for k, v in entry_to_gene.items() if v}

                        # entry_name → entry_name, matching uppercased
                        entry_name_upper_to_entry = {k.upper(): k for k in entry_to_gene.keys()}

                        # stripped (non-species-specific) → entry_name
                        entry_name_no_species_to_entry = {
                            k.split('_')[0].upper(): k for k in entry_to_gene.keys()
                        }

                        # Load excel file (workbook) and get sheet names #
                        sheet_names = workbook.sheetnames

                        # Sheets and headers #

                        Sheet_pass_check = False
                        valid_sheet_name_list = ['GPCRome wheel - Numeric','GPCRome wheel - Text','Tree - Numeric','Tree - Text','Cluster - Numeric','List - Numeric','List - Text','Heatmap - Numeric']

                        # Figure out if the sheet names match a template file #
                        for sheet_name in sheet_names:
                            if sheet_name in valid_sheet_name_list:
                                Sheet_pass_check = True
                            else:
                                pass

                        if not Sheet_pass_check:
                            # Add addition for the different sheets.
                            return render(request, f'mapper/DataMapperHome{self.page}.html', {'upload_status': 'Failed','Error_message': "The excel file is not structured as the template file. There are incorrect sheet names and data setup."})
                        else:

                            ### Initialize Global values ###
                            # Data passed #
                            Data = {}
                            # Incorrect values for the report #
                            Incorrect_values = {}
                            # plot evaluation value #
                            Plot_parser = 'Failed'
                            # Plot name & Datatype
                            plot_name_global = ""
                            plot_type_global = ""

                            def is_valid_color(color):
                                # Check if it's a named color
                                if color.lower() in mcolors.CSS4_COLORS:
                                    return True

                                # Check if it's a valid hex color
                                hex_pattern = r"^#([A-Fa-f0-9]{6}|[A-Fa-f0-9]{3})$"
                                return bool(re.match(hex_pattern, color))

                            def normalize_receptor_input(receptor_raw):
                                if not receptor_raw:
                                    return None

                                receptor = receptor_raw.strip().upper()

                                # Gene name match
                                if receptor in gene_to_entry:
                                    return gene_to_entry[receptor]

                                # Full UniProt-style match
                                if receptor in entry_name_upper_to_entry:
                                    return entry_name_upper_to_entry[receptor]

                                # Fallback: stripped version without species suffix
                                if receptor in entry_name_no_species_to_entry:
                                    return entry_name_no_species_to_entry[receptor]

                                # No match
                                return None

                            # For each sheet in the workbook #
                            for sheet_name in sheet_names:

                                if sheet_name in valid_sheet_name_list:
                                    Plot_name = sheet_name.split(" - ")[0]
                                    Plot_type = sheet_name.split(" - ")[1]

                                    # Set global plot name and datatype
                                    plot_name_global = Plot_name
                                    plot_type_global = Plot_type
                                    # Initialize worksheet #
                                    worksheet = workbook[sheet_name]

                                    ##########################
                                    ### GPCRome wheel Plot ###
                                    ##########################
                                    if Plot_name == "GPCRome wheel":
                                        # Set empty sheet value
                                        empty_sheet = True

                                        # Efficient empty check using values only
                                        for row in worksheet.iter_rows(min_row=2, values_only=True):
                                            if row[1] not in (None, ""):
                                                empty_sheet = False
                                                break

                                        if empty_sheet:
                                            pass #If sheet is empty pass the checking and report the findings
                                        else:
                                            for index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):

                                                # Get raw input
                                                receptor_raw = row[0].value

                                                # Skip empty rows early
                                                if receptor_raw in (None, ""):
                                                    continue

                                                # Normalize input
                                                receptor = normalize_receptor_input(receptor_raw)

                                                # Validate receptor
                                                if receptor not in Proteins_GPCRomeTree:
                                                    Incorrect_values.setdefault('GPCRs', {})[index] = f'"{row[0].value}" is an invalid entry for the GPCRome wheel plot.'
                                                    continue

                                                if receptor not in Data:
                                                    Data[receptor] = {}

                                                # ==== NUMERIC ====
                                                if Plot_type == 'Numeric':
                                                    DataValue = row[1].value
                                                    if DataValue not in (None, ""):
                                                        try:
                                                            Data[receptor]['Value1'] = float(DataValue)
                                                        except ValueError:
                                                            Incorrect_values.setdefault('GPCR Value (Column B)', {})[index] = 'Non-numeric Value'

                                                # ==== TEXT ====
                                                elif Plot_type == 'Text':
                                                    DataValue = row[1].value
                                                    color_fill_cell = row[2] if len(row) > 2 else None
                                                    color_override_cell = row[3] if len(row) > 3 else None

                                                    if DataValue not in (None, ""):
                                                        try:
                                                            Data[receptor]['Value1'] = str(DataValue)
                                                        except ValueError:
                                                            Incorrect_values.setdefault('GPCR Value (Column B)', {})[index] = 'Corrupted Value, not a string'

                                                    # Handle optional fill and override color
                                                    color_fill = None
                                                    color_override = None

                                                    if color_fill_cell and color_fill_cell.fill:
                                                        fill = color_fill_cell.fill
                                                        start_color = fill.start_color

                                                        if start_color:
                                                            # Case 1: Normal RGB fill (ARGB or RGB)
                                                            if start_color.type == 'rgb' and start_color.rgb:
                                                                argb = start_color.rgb
                                                                if len(argb) == 8:
                                                                    color_fill = f"#{argb[-6:]}"  # Strip alpha
                                                                elif len(argb) == 6:
                                                                    color_fill = f"#{argb}"

                                                            # Case 2: Theme fill – warn user
                                                            elif start_color.type == 'theme':
                                                                Incorrect_values.setdefault('GPCR Value (Column C)', {})[index] = (
                                                                    'Theme-based cell color detected. Please use standard RGB fill (not theme colors - or use "Optional: Assign hex codes" only).'
                                                                )

                                                    if color_override_cell and color_override_cell.value not in (None, ""):
                                                        color_override = str(color_override_cell.value).strip()

                                                    if color_override:
                                                        if is_valid_color(color_override):
                                                            Data[receptor]['Value2'] = color_override
                                                        else:
                                                            Incorrect_values.setdefault('GPCR Value (Column D)', {})[index] = 'Not a valid Hexcode'
                                                            Data[receptor]['Value2'] = "Black"
                                                    elif color_fill:
                                                        Data[receptor]['Value2'] = color_fill
                                                    else:
                                                        Data[receptor]['Value2'] = "Black"
                                                else:
                                                    Incorrect_values.setdefault('Errors', {})[index] = (
                                                        'Incorrect datatype: Unable to determine if Numeric or Text. Excel sheet may not match template.'
                                                    )

                                        # === Determine plot status ===
                                        status = 'Success'

                                        if empty_sheet:
                                            status = 'Empty sheet'
                                        elif Data:
                                            for field_errors in Incorrect_values.values():
                                                if any(field_errors.values()):
                                                    status = 'Failed'
                                                    break
                                        else:
                                            status = 'Failed'

                                        Plot_parser = status
                            
                                    #################
                                    ### Tree plot ###
                                    #################

                                    elif Plot_name == 'Tree':
                                        
                                        # Incorrect_values['GPCRs'] = {}
                                        IndexToColumn = ['A','B','C','D','E','F','G']
                                        empty_sheet = True  # Initialize the flag
                                        TreeNonEmptyEntryCounter = 0
                                        
                                        # Check if sheet is empty
                                        for row in worksheet.iter_rows(min_row=2, values_only=True):
                                            # Check if any value in columns B (1) to G (6) is not None or empty
                                            if any(cell not in (None, "") for cell in row[1:7]):  # Slice [1:7] to include columns B-G
                                                empty_sheet = False
                                                break
                                        # If sheet is empty continue and report
                                        if empty_sheet:
                                            pass
                                        else:
                                            # Iterate over rows starting from row 2 (min_row=2)
                                            for row in worksheet.iter_rows(min_row=2, values_only=True):
                                                # Check if column A has a value AND at least one column in B-G has a value
                                                if row[0] not in (None, "") and any(cell not in (None, "") for cell in row[1:7]):
                                                    TreeNonEmptyEntryCounter += 1
                                                
                                                # Stop if we exceed 200 valid entries
                                                if TreeNonEmptyEntryCounter > 200:
                                                    break

                                            if TreeNonEmptyEntryCounter > 200:
                                                Incorrect_values['Error'] = "Detected more than 200 entries. The Tree can not handle that many GPCRs to be able to visualize them in one plot."
                                            else:
                                                # If sheet is not empty then start handling the data
                                                # Iterate through the rows and validate the data
                                                for index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):

                                                    # Get raw input
                                                    receptor_raw = row[0].value

                                                    # Skip empty rows early
                                                    if receptor_raw in (None, ""):
                                                        continue

                                                    # Normalize input
                                                    receptor = normalize_receptor_input(receptor_raw)

                                                    # Check the "Receptor (Uniprot)" column for correct values
                                                    if receptor not in Proteins_GPCRomeTree:
                                                        if receptor in (None, ""):
                                                            pass
                                                        else:
                                                            Incorrect_values['GPCRs'][index] = '"{}" is a invalid entry for the Tree plot.'.format(receptor)
                                                    else:
                                                        # Check if data row is in data and/or initialize it
                                                        if receptor not in Data:
                                                            Data[receptor] = {}
                                                        # Check datatype -> Numeric:
                                                        if Plot_type  == 'Numeric':
                                                            for i in range(1,7):

                                                                if row[i].value not in (None, ""):
                                                                    try:
                                                                        float_value = float(row[i].value)
                                                                        if i == 1:
                                                                            Data[receptor]['Inner'] = float_value
                                                                        else:
                                                                            Data[receptor]['Outer{}'.format(i-1)] = float_value
                                                                    except ValueError:
                                                                        column_key = 'GPCR Value (Column {})'.format(IndexToColumn[i])
                                                                        Incorrect_values.setdefault(column_key, {})[index] = 'Non-numeric Value'
                                                        # ==== TEXT ====
                                                        elif Plot_type == 'Text':
                                                            DataValue = row[1].value
                                                            color_fill_cell = row[2] if len(row) > 2 else None
                                                            color_override_cell = row[3] if len(row) > 3 else None

                                                            if DataValue not in (None, ""):
                                                                try:
                                                                    Data[receptor]['Inner'] = str(DataValue)
                                                                except ValueError:
                                                                    Incorrect_values.setdefault('GPCR Value (Column B)', {})[index] = 'Corrupted Value, not a string'

                                                            # Handle optional fill and override color
                                                            color_fill = None
                                                            color_override = None

                                                            if color_fill_cell and color_fill_cell.fill:
                                                                fill = color_fill_cell.fill
                                                                start_color = fill.start_color

                                                                if start_color:
                                                                    # Case 1: Normal RGB fill (ARGB or RGB)
                                                                    if start_color.type == 'rgb' and start_color.rgb:
                                                                        argb = start_color.rgb
                                                                        if len(argb) == 8:
                                                                            color_fill = f"#{argb[-6:]}"  # Strip alpha
                                                                        elif len(argb) == 6:
                                                                            color_fill = f"#{argb}"

                                                                    # Case 2: Theme-based fill – show warning
                                                                    elif start_color.type == 'theme':
                                                                        Incorrect_values.setdefault('GPCR Value (Column C)', {})[index] = (
                                                                            'Theme-based cell color detected. Please use standard RGB fill (not theme colors - or use "Optional: Assign hex codes" only).'
                                                                        )

                                                            if color_override_cell and color_override_cell.value not in (None, ""):
                                                                color_override = str(color_override_cell.value).strip()

                                                            if color_override:
                                                                if is_valid_color(color_override):
                                                                    Data[receptor]['ColorValue'] = color_override
                                                                else:
                                                                    Incorrect_values.setdefault('GPCR Value (Column D)', {})[index] = 'Not a valid Hexcode'
                                                                    Data[receptor]['ColorValue'] = "Black"
                                                            elif color_fill:
                                                                Data[receptor]['ColorValue'] = color_fill
                                                            else:
                                                                Data[receptor]['ColorValue'] = "Black"
                                                        else:
                                                            Incorrect_values['Errors'] = 'Incorrect datatype: Was unable to determine datatype to be either Numeric or text. Template might have been changed to something that the DataMapper can not handle.'
                                        # Check if any values are incorrect #
                                        status = 'Success'

                                        if empty_sheet:
                                            status = 'Empty sheet'
                                        elif Data:
                                            for Error_entries in Incorrect_values:
                                                # Check if there are any assigned index values for this col_idx
                                                if any(Incorrect_values[Error_entries].values()):
                                                    # If any index is assigned, set status to 'Partially_success' and break out of the loop
                                                    status = 'Failed'
                                                    break
                                        else:
                                            status = 'Failed'

                                        ## Update Plot_parser for T
                                        Plot_parser = status

                                    ####################
                                    ### Cluster plot ###
                                    ####################

                                    elif Plot_name == 'Cluster':

                                        ClusterSpacialPositionCheck = False
                                        empty_sheet = True  # Initialize the flag

                                        # Check if the sheet is empty
                                        for row in worksheet.iter_rows(min_row=2, values_only=True):
                                            # Check if any value in columns B (1) and C (2) is not None or empty
                                            if any(cell not in (None, "") for cell in (row[1], row[2])):  # Only check B and C
                                                empty_sheet = False
                                                break

                                        # If sheet is empty continue and report
                                        if empty_sheet:
                                            pass
                                        else:
                                            # Iterate over rows starting from row 2
                                            for row in worksheet.iter_rows(min_row=2, values_only=True):
                                                # Check if both column A (0) and column C (2) have values
                                                if row[0] not in (None, "") and row[2] not in (None, ""):
                                                    ClusterSpacialPositionCheck = True
                                                    break  # Stop checking as soon as we find a valid entry

                                            # If sheet is not empty then start handling the data
                                            # Iterate through the rows and validate the data
                                            for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):

                                                # Get raw input
                                                receptor_raw = row[0]

                                                # Skip empty rows early
                                                if receptor_raw in (None, ""):
                                                    continue

                                                # Normalize input
                                                receptor = normalize_receptor_input(receptor_raw)

                                                if receptor_raw == "RRH":
                                                    print(receptor_raw, receptor)
                                                # Validate receptor
                                                if receptor not in all_proteins:
                                                    # Incorrect_values['GPCRs'][index] = '"{}" is a invalid entry for the Cluster plot.'.format(row[0])
                                                    Incorrect_values.setdefault('GPCRs', {})[index] = '"{}" is a invalid entry for the Cluster plot.'.format(row[0])
                                                    continue

                                                # Check if data row is in data and/or initialize it
                                                if receptor not in Data:
                                                    Data[receptor] = {}

                                                if ClusterSpacialPositionCheck:
                                                    # Check datatype -> Numeric
                                                    if Plot_type == 'Numeric':
                                                        if row[1] not in (None, ""):
                                                            try:
                                                                float_value = float(row[1])
                                                                Data[receptor]['Value1'] = float_value
                                                            except ValueError:
                                                                Incorrect_values.setdefault('GPCR Value (Column B)', {})[index] = 'Non-numeric Value.'
                                                            
                                                            if row[2] not in (None, ""):
                                                                try:
                                                                    float_value = float(row[2])
                                                                    Data[receptor]['Value2'] = float_value
                                                                except ValueError:
                                                                    Incorrect_values.setdefault('GPCR Value (Column C)', {})[index] = 'Non-numeric Value.'
                                                        else:
                                                            if row[2] not in (None, ""):
                                                                Incorrect_values.setdefault('GPCR Value (Column B & C)', {})[index] = 'You have Spacial position (Column C) without a value assign (Column B), Please assign a value to this row in Column B.'
                                                    else:
                                                        Incorrect_values['Errors'] = 'Incorrect datatype: Was unable to determine datatype to be Numeric. Template might have been changed to something that the DataMapper can not handle.'
                                                else:
                                                    # Check datatype -> Numeric
                                                    if Plot_type == 'Numeric':
                                                        if row[1] not in (None, ""):
                                                            try:
                                                                float_value = float(row[1])
                                                                Data[receptor]['Value1'] = float_value
                                                            except ValueError:
                                                                Incorrect_values.setdefault('GPCR Value (Column B)', {})[index] = 'Non-numeric Value.'
                                                    else:
                                                        Incorrect_values['Errors'] = 'Incorrect datatype: Was unable to determine datatype to be Numeric. Template might have been changed to something that the DataMapper can not handle.'

                                        # Check if any values are incorrect #
                                        status = 'Success'

                                        if empty_sheet:
                                            status = 'Empty sheet'
                                        elif Data:
                                            for Error_entries in Incorrect_values:
                                                # Check if there are any assigned index values for this col_idx
                                                if any(Incorrect_values[Error_entries].values()):
                                                    # If any index is assigned, set status to 'Failed' and break out of the loop
                                                    status = 'Failed'
                                                    break
                                        else:
                                            status = 'Failed'
                                            
                                        ## Update Plot_parser for Cluster
                                        Plot_parser = status


                                    #################
                                    ### List plot ###
                                    #################

                                    elif Plot_name == 'List':

                                        IndexToColumn = ['A','B','C','D','E','F','G',"H","I"]
                                        # # Initialize the inccorect column values
                                        empty_sheet = True  # Initialize the flag
                                        
                                        # Check if sheet is empty
                                        for row in worksheet.iter_rows(min_row=2, values_only=True):
                                            # Check if any value in columns B (1) to G (6) is not None or empty
                                            if any(cell not in (None, "") for cell in row[1:7]):  # Slice [1:7] to include columns B-G
                                                empty_sheet = False
                                                break

                                        # If sheet is empty continue and report
                                        if empty_sheet:
                                            pass
                                        else:
                                            # If sheet is not empty then start handling the data
                                            # Iterate through the rows and validate the data
                                            for index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):

                                                 # Get raw input
                                                receptor_raw = row[0].value

                                                # Skip empty rows early
                                                if receptor_raw in (None, ""):
                                                    continue

                                                # Normalize input
                                                receptor = normalize_receptor_input(receptor_raw)

                                                # Validate receptor
                                                if receptor not in all_proteins:
                                                    Incorrect_values.setdefault('GPCRs', {})[index] = f'"{receptor}" is an invalid entry for the  List plot.'
                                                    continue

                                                if receptor not in Data:
                                                    Data[receptor] = {}

                                                # ==== NUMERIC ====
                                                if Plot_type == 'Numeric':
                                                    for i in range(1, 5):
                                                        cell_value = row[i].value
                                                        col_label = 'GPCR Value (Column {})'.format(IndexToColumn[i])

                                                        if cell_value not in (None, ""):
                                                            if Plot_type == 'Numeric':
                                                                try:
                                                                    float_value = float(cell_value)
                                                                    Data.setdefault(receptor, {})['Value{}'.format(i)] = float_value
                                                                except ValueError:
                                                                    Incorrect_values.setdefault(col_label, {})[index] = 'Non-numeric Value'

                                                            elif Plot_type == 'Text':
                                                                try:
                                                                    text_value = str(cell_value)
                                                                    Data.setdefault(receptor, {})['Value{}'.format(i)] = text_value
                                                                except ValueError:
                                                                    Incorrect_values.setdefault(col_label, {})[index] = 'Corrupted Value, not a string'

                                                    # Handle shapes in columns F (5) to I (8)
                                                    for i in range(5, 9):
                                                        shape_value = row[i].value
                                                        col_label = 'GPCR Value (Column {})'.format(IndexToColumn[i])

                                                        if shape_value not in (None, ""):
                                                            if shape_value in ('Circle', 'Diamond', 'Rectangle', 'Star', 'Triangle'):
                                                                Data.setdefault(receptor, {})['Value{}'.format(i)] = str(shape_value)
                                                            else:
                                                                Incorrect_values.setdefault(col_label, {})[index] = (
                                                                    'Value assigned is not empty, Circle, Diamond, Rectangle, Star, or Triangle.'
                                                                )
                                                        else:
                                                            # If a shape is missing but its associated numeric/text value (B–E) is present
                                                            if row[i - 4].value not in (None, ""):
                                                                Data.setdefault(receptor, {})['Value{}'.format(i)] = 'Circle'
                                                # ==== TEXT ====
                                                elif Plot_type == 'Text':
                                                    DataValue = row[1].value if hasattr(row[1], 'value') else row[1]
                                                    color_fill_cell = row[2] if len(row) > 2 else None
                                                    color_override_cell = row[3] if len(row) > 3 else None

                                                    if DataValue not in (None, ""):
                                                        try:
                                                            Data[receptor]['Inner'] = str(DataValue)
                                                        except Exception:
                                                            Incorrect_values.setdefault('GPCR Value (Column B)', {})[index] = 'Corrupted Value, not a string'

                                                    # Handle optional fill and override color
                                                    color_fill = None
                                                    color_override = None

                                                    if color_fill_cell and hasattr(color_fill_cell, 'fill'):
                                                        fill = color_fill_cell.fill
                                                        start_color = getattr(fill, 'start_color', None)

                                                        if start_color:
                                                            # Case 1: Normal RGB fill (ARGB or RGB)
                                                            if start_color.type == 'rgb' and start_color.rgb:
                                                                argb = start_color.rgb
                                                                if len(argb) == 8:  # e.g. 'FF112233'
                                                                    color_fill = f"#{argb[-6:]}"
                                                                elif len(argb) == 6:  # e.g. '112233'
                                                                    color_fill = f"#{argb}"
                                                            # Case 2: Theme-based fill
                                                            elif start_color.type == 'theme':
                                                                Incorrect_values.setdefault('GPCR Value (Column C)', {})[index] = (
                                                                    'Theme-based cell color detected. Please use standard RGB fill (not theme colors - or use "Optional: Assign hex codes" only).'
                                                                )

                                                    if color_override_cell and getattr(color_override_cell, 'value', None) not in (None, ""):
                                                        color_override = str(color_override_cell.value).strip()

                                                    if color_override:
                                                        if is_valid_color(color_override):
                                                            Data[receptor]['ColorValue'] = color_override
                                                        else:
                                                            Incorrect_values.setdefault('GPCR Value (Column D)', {})[index] = 'Not a valid Hexcode'
                                                            Data[receptor]['ColorValue'] = "Black"
                                                    elif color_fill:
                                                        Data[receptor]['ColorValue'] = color_fill
                                                    else:
                                                        Data[receptor]['ColorValue'] = "Black"
                                                else:
                                                    Incorrect_values['Errors'] = 'Incorrect datatype: Was unable to determine datatype to be either Numeric or Text. Template might have been changed to something that the DataMapper cannot handle.'

                                        # Check if any values are incorrect #
                                        status = 'Success'

                                        if empty_sheet:
                                            status = 'Empty sheet'
                                        elif Data:
                                            for Error_entries in Incorrect_values:
                                                # Check if there are any assigned index values for this col_idx
                                                if any(Incorrect_values[Error_entries].values()):
                                                    # If any index is assigned, set status to 'Partially_success' and break out of the loop
                                                    status = 'Failed'
                                                    break
                                        else:
                                            status = 'Failed'

                                        ## Update Plot_parser for GPCRome Wheel
                                        Plot_parser = status

                                    ###############
                                    ### Heatmap ###
                                    ###############

                                    elif Plot_name == 'Heatmap':
                                        ## Initialize Local values ##
                                        IndexToColumn = ['A','B','C','D','E','F']
                                        # Initialize the inccorect column values
                                        empty_sheet = True  # Initialize the flag
                                        HeatmapNonEmptyEntryCounter = 0

                                        # Check if sheet is empty
                                        for row in worksheet.iter_rows(min_row=2, values_only=True):
                                            # Check if any value in columns B (1) to G (6) is not None or empty
                                            if any(cell not in (None, "") for cell in row[1:6]):  # Slice [1:6] to include columns B-F
                                                empty_sheet = False
                                                break

                                        # If sheet is empty continue and report
                                        if empty_sheet:
                                            pass
                                        else:
                                            # Iterate over rows starting from row 2 (min_row=2)
                                            for row in worksheet.iter_rows(min_row=2, values_only=True):
                                                # Check if column A has a value AND at least one column in B-G has a value
                                                if row[0] not in (None, "") and any(cell not in (None, "") for cell in row[1:7]):
                                                    HeatmapNonEmptyEntryCounter += 1

                                                # Stop if we exceed 200 valid entries
                                                if HeatmapNonEmptyEntryCounter > 50:
                                                    break

                                            if HeatmapNonEmptyEntryCounter > 50:
                                                Incorrect_values['Error'] = "Detected more than 50 entries. The Heatmap can not handle that many GPCRs to be able to visualize them in one plot."
                                            else:
                                                # If sheet is not empty then start handling the data
                                                # Iterate through the rows and validate the data
                                                for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):

                                                     # Get raw input
                                                    receptor_raw = row[0]

                                                    # Skip empty rows early
                                                    if receptor_raw in (None, ""):
                                                        continue

                                                    # Normalize input
                                                    receptor = normalize_receptor_input(receptor_raw)

                                                    # Validate receptor
                                                    if receptor not in all_proteins:
                                                        Incorrect_values.setdefault('GPCRs', {})[index] = f'"{receptor}" is an invalid entry for the Heatmap plot.'
                                                        continue

                                                    if receptor not in Data:
                                                        Data[receptor] = {}
                                                    
                                                    # Check datatype -> Numeric or Text:
                                                    if Plot_type in ('Numeric', 'Text'):
                                                        for i in range(1, 6):
                                                            cell_value = row[i]
                                                            col_label = 'GPCR Value (Column {})'.format(IndexToColumn[i])

                                                            if Plot_type == 'Numeric':
                                                                if cell_value not in (None, ""):
                                                                    try:
                                                                        float_value = float(cell_value)
                                                                        Data.setdefault(receptor, {})['Value{}'.format(i)] = float_value
                                                                    except ValueError:
                                                                        Incorrect_values.setdefault(col_label, {})[index] = 'Non-numeric Value'
                                                    else:
                                                        Incorrect_values['Errors'] = 'Incorrect datatype: Was unable to determine datatype to be either Numeric or text. Template might have been changed to something that the DataMapper can not handle.'
                                                # Check if any values are incorrect #
                                                status = 'Success'

                                                if empty_sheet:
                                                    status = 'Empty sheet'
                                                elif Data:
                                                    for Error_entries in Incorrect_values:
                                                        # Check if there are any assigned index values for this col_idx
                                                        if any(Incorrect_values[Error_entries].values()):
                                                            # If any index is assigned, set status to 'Partially_success' and break out of the loop
                                                            status = 'Failed'
                                                            break
                                                else:
                                                    status = 'Failed'

                                                ## Update Plot_parser for GPCRome Wheel
                                                Plot_parser = status

                            #######################################################
                            ## Evaluate the uploaded data and pass on the report ##
                            #######################################################

                            plot_data_json = json.dumps(Data) if Data else "No Data"

                            if Plot_parser in ('Success','Empty sheet'):
                                context = {'upload_status': 'Success',
                                        'Plot_parser': Plot_parser,
                                        'plot_name': plot_name_global,
                                        'plot_type': plot_type_global,
                                        'Data': plot_data_json}
                            else:
                                context = {'upload_status': 'Success',
                                        'Plot_parser': Plot_parser,
                                        'plot_name': plot_name_global,
                                        'plot_type': plot_type_global,
                                        'Data': plot_data_json}
                                if Incorrect_values:
                                    context['Incorrect_data_json'] = Incorrect_values
                                else:
                                    context['Incorrect_data_json'] = "No incorrect data"

                            # Return the rendered results
                            return render(request, f'mapper/DataMapperHome{self.page}.html', context)

                    else:
                        return render(request, f'mapper/DataMapperHome{self.page}.html', {'upload_status': 'Failed','Error_message': "Unable to load excel file, might be corrupted or not inline with a template file."})

            else:
                # Return a 405 Method Not Allowed response if not a POST request
                return render(request, f'mapper/DataMapperHome{self.page}.html', {'upload_status': 'Failed','Error_message': "Not a valid excel file. Please try and use the template excel file."})

#######################
## Excel upload form ##
#######################

class ExcelUploadForm(forms.Form):
    file = forms.FileField()

class GPCRomeRender(TemplateView):
    template_name = 'mapper/PlotRender_GPCRome.html'  # default fallback

    def get_template_names(self):
        # If PlotType is set during post, switch template accordingly
        plot_type = getattr(self, 'plot_type', None)
        if plot_type == "Text":
            return ['mapper/PlotRender_GPCRome_Text.html']
        return ['mapper/PlotRender_GPCRome.html']

    def post(self, request, *args, **kwargs):
        Data_json = request.POST.get('Data')
        PlotType = request.POST.get('PlotType')

        try:
            Data = json.loads(Data_json)
            gpcr_data = DataMapperHome.GenerateGPCRomeDataStructure(type="Classic")
            updated_data = DataMapperHome.update_nested_GPCRome_data(gpcr_data["Data"], Data)

            # Store PlotType so get_template_names can access it
            self.plot_type = PlotType

            context = {
                'GPCRomeData': updated_data,
                'PlotType': PlotType
            }
            return self.render_to_response(context)

        except json.JSONDecodeError:
            return HttpResponse("Invalid JSON data")

class TreeRender(TemplateView):
    template_name = 'mapper/PlotRender_Tree.html'  # default fallback

    def get_template_names(self):
        # If PlotType is set during post, switch template accordingly
        plot_type = getattr(self, 'plot_type', None)
        if plot_type == "Text":
            return ['mapper/PlotRender_Tree_Text.html']
        return ['mapper/PlotRender_Tree.html']

    def post(self, request, *args, **kwargs):
        Data_json = request.POST.get('Data')
        PlotType = request.POST.get('PlotType')

        try:
            Data = json.loads(Data_json)
            tree, tree_options, circles, receptors, genes = DataMapperHome.generate_tree_plot(Data)

            # Store PlotType so get_template_names can access it
            self.plot_type = PlotType

            context = {
                'tree': json.dumps(tree),
                'tree_options': tree_options,
                'circles': json.dumps(circles),
                'Receptor_dict': json.dumps(receptors),
                'Entrez_dict': json.dumps(genes),
                'PlotType': PlotType,
                'Data': json.dumps(Data)
            }
            return self.render_to_response(context)

        except json.JSONDecodeError:
            return HttpResponse("Invalid JSON data")
        
class ClusterRender(TemplateView):
    template_name = 'mapper/PlotRender_Cluster.html'  # default fallback

    def post(self, request, *args, **kwargs):
        Data_json = request.POST.get('Data')
        # PlotType = request.POST.get('PlotType')

        try:
            # Get data
            Data = json.loads(Data_json)
            # Calculate the plot
            output_seq = DataMapperHome.clustering_test('tsne', Data,'seq')
            label_converter = DataMapperHome.Label_conversion_info(Data)
            # Create the context
            context = {
                'cluster_data_seq': output_seq,
                'Label_converter': json.dumps(label_converter)
                }
            # Return context
            return self.render_to_response(context)

        except json.JSONDecodeError:
            return HttpResponse("Invalid JSON data")
        
class ListRender(TemplateView):
    template_name = 'mapper/PlotRender_List.html'  # default fallback

    def get_template_names(self):
        # If PlotType is set during post, switch template accordingly
        plot_type = getattr(self, 'plot_type', None)
        if plot_type == "Text":
            return ['mapper/PlotRender_List_Text.html']
        return ['mapper/PlotRender_List.html']

    def post(self, request, *args, **kwargs):
        Data_json = request.POST.get('Data')
        PlotType = request.POST.get('PlotType')

        try:
            Data = json.loads(Data_json)
            listplot_data = DataMapperHome.generate_list_plot(Data)
            label_converter = DataMapperHome.Label_conversion_info(Data)

            # Store PlotType so get_template_names can access it
            self.plot_type = PlotType

            context = {
                'listplot_data': json.dumps(listplot_data["NameList"]),
                'listplot_data_variables': json.dumps(listplot_data['DataPoints']),
                'Label_Conversion': json.dumps(label_converter),
                'PlotType': PlotType
            }
            return self.render_to_response(context)

        except json.JSONDecodeError:
            return HttpResponse("Invalid JSON data")

class HeatmapRender(TemplateView):
    template_name = 'mapper/PlotRender_Heatmap.html'  # default fallback

    def post(self, request, *args, **kwargs):
        Data_json = request.POST.get('Data')
        # PlotType = request.POST.get('PlotType')

        try:
            # Get data
            Data = json.loads(Data_json)
            # Calculate the plot
            label_converter = DataMapperHome.Label_conversion_info(Data)
            # Create the context
            context = {
                'Label_converter': json.dumps(label_converter),
                'heatmap_data': json.dumps(Data)
                }
            # Return context
            return self.render_to_response(context)

        except json.JSONDecodeError:
            return HttpResponse("Invalid JSON data")